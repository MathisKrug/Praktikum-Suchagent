"""Erzeugt das mobile Dashboard (docs/index.html) und die Excel-Datei."""

from __future__ import annotations

import html
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NEW_WINDOW_HOURS = 36  # was in diesem Zeitraum zuerst gesehen wurde, gilt als "neu"


def _is_new(first_seen: str) -> bool:
    try:
        fs = datetime.fromisoformat(first_seen)
        if fs.tzinfo is None:
            fs = fs.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - fs).total_seconds() < NEW_WINDOW_HOURS * 3600
    except Exception:
        return False


def render_html(jobs: list[dict], errors: list[str], out: str = "docs/index.html",
                employers: list[dict] | None = None) -> None:
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    employers = employers or []

    payload = []
    for j in jobs:
        payload.append({
            "company": j["company"],
            "sector": j.get("sector") or "",
            "title": j["title"],
            "url": j["url"],
            "location": j.get("location") or "",
            "score": j["score"],
            "duration": j.get("duration"),
            "start": j.get("start_date") or "",
            "reasons": j.get("reasons") or "",
            "new": _is_new(j.get("first_seen") or ""),
            "first_seen": (j.get("first_seen") or "")[:10],
        })

    emp_payload = [{
        "name": e["name"],
        "hits": e["hits"],
        "best": e["best_score"],
        "sample": e.get("sample_title") or "",
        "watched": bool(e.get("promoted")),
        "adapter": e.get("adapter") or "",
    } for e in employers]

    new_count = sum(1 for p in payload if p["new"])
    watched_count = sum(1 for e in emp_payload if e["watched"])
    err_block = ""
    if errors:
        items = "".join(f"<li>{html.escape(e)}</li>" for e in errors)
        err_block = (
            f'<details class="err"><summary>{len(errors)} Quelle(n) mit Problemen</summary>'
            f"<ul>{items}</ul></details>"
        )

    doc = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="Praktikum-Radar">
<meta name="theme-color" content="#111418">
<title>Praktikum-Radar</title>
<style>
  :root {{ --bg:#0f1115; --card:#181b21; --line:#262b33; --tx:#e8eaed;
           --dim:#9aa2ad; --acc:#4da3ff; --new:#2ecc71; }}
  * {{ box-sizing:border-box; -webkit-tap-highlight-color:transparent; }}
  body {{ margin:0; background:var(--bg); color:var(--tx);
          font:15px/1.45 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
          padding:env(safe-area-inset-top) 0 env(safe-area-inset-bottom); }}
  header {{ padding:20px 16px 12px; position:sticky; top:0; background:var(--bg);
            border-bottom:1px solid var(--line); z-index:10; }}
  h1 {{ margin:0 0 4px; font-size:20px; letter-spacing:-.3px; }}
  .meta {{ color:var(--dim); font-size:12px; }}
  .bar {{ display:flex; gap:6px; padding:12px 16px; overflow-x:auto;
          border-bottom:1px solid var(--line); }}
  .bar button {{ background:var(--card); color:var(--tx); border:1px solid var(--line);
                 border-radius:999px; padding:7px 13px; font-size:13px; white-space:nowrap; }}
  .bar button.on {{ background:var(--acc); border-color:var(--acc); color:#04121f; font-weight:600; }}
  main {{ padding:12px 16px 40px; }}
  .job {{ background:var(--card); border:1px solid var(--line); border-radius:12px;
          padding:14px; margin-bottom:10px; display:block; color:inherit;
          text-decoration:none; }}
  .job:active {{ border-color:var(--acc); }}
  .top {{ display:flex; justify-content:space-between; gap:10px; align-items:flex-start; }}
  .co {{ font-size:12px; color:var(--dim); text-transform:uppercase; letter-spacing:.4px; }}
  .ti {{ font-weight:600; margin:3px 0 6px; }}
  .sc {{ background:#20252d; border-radius:8px; padding:3px 9px; font-size:13px;
         font-weight:700; color:var(--acc); flex-shrink:0; }}
  .tags {{ display:flex; flex-wrap:wrap; gap:5px; margin-top:8px; }}
  .tag {{ font-size:11px; color:var(--dim); background:#20252d; padding:3px 8px;
          border-radius:6px; }}
  .badge {{ background:var(--new); color:#05210f; font-weight:700; }}
  .why {{ font-size:11px; color:var(--dim); margin-top:8px; line-height:1.5; }}
  .err {{ margin:14px 16px; font-size:12px; color:var(--dim); }}
  .err summary {{ cursor:pointer; }}
  .empty {{ color:var(--dim); text-align:center; padding:50px 20px; line-height:1.7; }}
  .tabs {{ display:flex; gap:8px; margin-top:12px; }}
  .tabs button {{ background:none; border:none; border-bottom:2px solid transparent;
                  color:var(--dim); font-size:14px; padding:6px 2px; font-weight:600; }}
  .tabs button.on {{ color:var(--tx); border-bottom-color:var(--acc); }}
  .note {{ font-size:12px; color:var(--dim); margin-bottom:12px; line-height:1.6; }}
</style>
</head>
<body>
<header>
  <h1>Praktikum-Radar</h1>
  <div class="meta">{len(payload)} Treffer &middot; {new_count} neu &middot;
    {len(emp_payload)} Firmen entdeckt &middot; Stand {now}</div>
  <div class="tabs">
    <button data-v="jobs" class="on">Stellen</button>
    <button data-v="firmen">Firmen ({len(emp_payload)})</button>
  </div>
</header>
<div class="bar" id="filters">
  <button data-f="all" class="on">Alle</button>
  <button data-f="new">Neu</button>
  <button data-f="fashion">Fashion</button>
  <button data-f="fmcg">FMCG</button>
  <button data-f="entdeckt">Entdeckt</button>
  <button data-f="short">3&ndash;4 Monate</button>
</div>
{err_block}
<main id="list"></main>
<script>
const JOBS = {json.dumps(payload, ensure_ascii=False)};
const EMPLOYERS = {json.dumps(emp_payload, ensure_ascii=False)};
const WATCHED = {watched_count};
const list = document.getElementById('list');
const filters = document.getElementById('filters');
let view = 'jobs', activeFilter = 'all';

function esc(s) {{
  return String(s == null ? '' : s).replace(/[&<>"]/g, c =>
    ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}})[c]);
}}

function render(filter) {{
  let rows = JOBS;
  if (filter === 'new') rows = rows.filter(j => j.new);
  else if (filter === 'short') rows = rows.filter(j => j.duration && j.duration <= 4.4);
  else if (filter !== 'all') rows = rows.filter(j => j.sector === filter);

  if (!rows.length) {{
    list.innerHTML = '<div class="empty">Nichts gefunden.</div>';
    return;
  }}

  list.innerHTML = rows.map(j => {{
    const tags = [];
    if (j.new) tags.push('<span class="tag badge">NEU</span>');
    if (j.location) tags.push('<span class="tag">' + esc(j.location) + '</span>');
    if (j.duration) tags.push('<span class="tag">' + j.duration + ' Mon.</span>');
    if (j.start) tags.push('<span class="tag">ab ' + esc(j.start) + '</span>');
    return '<a class="job" href="' + esc(j.url) + '" target="_blank" rel="noopener">'
      + '<div class="top"><div><div class="co">' + esc(j.company) + '</div>'
      + '<div class="ti">' + esc(j.title) + '</div></div>'
      + '<div class="sc">' + j.score + '</div></div>'
      + '<div class="tags">' + tags.join('') + '</div>'
      + '<div class="why">' + esc(j.reasons) + '</div></a>';
  }}).join('');
}}

function renderEmployers() {{
  if (!EMPLOYERS.length) {{
    list.innerHTML = '<div class="empty">Noch keine Firmen entdeckt.<br>'
      + 'Stufe 1 braucht API-Keys &ndash; siehe README.</div>';
    return;
  }}
  list.innerHTML = '<div class="note">' + EMPLOYERS.length + ' Arbeitgeber haben '
    + 'passende Praktika ausgeschrieben. ' + WATCHED + ' davon werden inzwischen '
    + 'direkt auf ihrer Karriereseite &uuml;berwacht.</div>'
    + EMPLOYERS.map(e => {{
    const tags = [];
    if (e.watched) tags.push('<span class="tag badge">&Uuml;BERWACHT</span>');
    tags.push('<span class="tag">' + e.hits + ' Treffer</span>');
    if (e.adapter) tags.push('<span class="tag">' + esc(e.adapter) + '</span>');
    return '<div class="job"><div class="top"><div>'
      + '<div class="ti">' + esc(e.name) + '</div>'
      + '<div class="co" style="text-transform:none">' + esc(e.sample) + '</div></div>'
      + '<div class="sc">' + e.best + '</div></div>'
      + '<div class="tags">' + tags.join('') + '</div></div>';
  }}).join('');
}}

function draw() {{
  filters.style.display = (view === 'jobs') ? 'flex' : 'none';
  if (view === 'jobs') render(activeFilter); else renderEmployers();
}}

document.querySelectorAll('.bar button').forEach(b => {{
  b.addEventListener('click', () => {{
    document.querySelectorAll('.bar button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    activeFilter = b.dataset.f;
    draw();
  }});
}});

document.querySelectorAll('.tabs button').forEach(b => {{
  b.addEventListener('click', () => {{
    document.querySelectorAll('.tabs button').forEach(x => x.classList.remove('on'));
    b.classList.add('on');
    view = b.dataset.v;
    draw();
  }});
}});

draw();
</script>
</body>
</html>"""

    Path(out).write_text(doc, encoding="utf-8")


def render_xlsx(jobs: list[dict], out: str = "docs/praktika.xlsx",
                employers: list[dict] | None = None) -> None:
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Treffer"

    headers = ["Neu", "Score", "Unternehmen", "Rolle", "Standort", "Monate",
               "Start", "Sektor", "Zuerst gesehen", "Status", "Notiz", "Link"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F3864")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")

    new_fill = PatternFill("solid", fgColor="D8F3DC")

    for j in jobs:
        is_new = _is_new(j.get("first_seen") or "")
        ws.append([
            "NEU" if is_new else "",
            j["score"],
            j["company"],
            j["title"],
            j.get("location") or "",
            j.get("duration") or "",
            j.get("start_date") or "",
            j.get("sector") or "",
            (j.get("first_seen") or "")[:10],
            j.get("status") or "offen",
            "",
            j["url"],
        ])
        if is_new:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = new_fill

    widths = [6, 7, 20, 52, 22, 8, 12, 12, 14, 12, 30, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"

    # Zweites Blatt: das entdeckte Firmenregister
    ws2 = wb.create_sheet("Firmen entdeckt")
    h2 = ["Unternehmen", "Treffer", "Bester Score", "Beispielrolle",
          "Wird überwacht", "System", "Zuerst gesehen"]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill

    for e in (employers or []):
        ws2.append([
            e.get("name", ""),
            e.get("hits", 0),
            e.get("best_score", 0),
            e.get("sample_title", ""),
            "ja" if e.get("promoted") else "",
            e.get("adapter") or "",
            (e.get("first_seen") or "")[:10],
        ])

    for i, w in enumerate([34, 9, 13, 50, 15, 17, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    if ws2.max_row > 1:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}{ws2.max_row}"

    wb.save(out)
